"""Brazilian retail fuel prices (BRL per litre).

The ANP publishes a weekly price survey. Its workbook carries a BRASIL
sheet holding official national averages, so no averaging is done here.

Two details shape the parsing:

  * A UNIDADE DE MEDIDA column states each row's unit. Only "R$/l" rows are
    read, which excludes LPG (sold per 13kg cylinder) and CNG (per cubic
    metre) - storing either as a litre price would be nonsense.
  * Product names are matched exactly rather than by substring, because
    "OLEO DIESEL S10" contains "OLEO DIESEL"; a substring match would file
    the low-sulphur grade as ordinary diesel.

Headers are accented, so text is compared with accents stripped.
"""

import io
import logging
import unicodedata
from datetime import datetime, timezone

import openpyxl
from bs4 import BeautifulSoup

from ..models import LocalPrice
from . import http

log = logging.getLogger(__name__)

ANP_PAGE_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas"
)
ANP_BASE = "https://www.gov.br"

# The weekly summary workbook; the "revendas" file is per-station detail.
LINK_MARKER = "resumo_semanal_lpc"

NATIONAL_SHEET = "BRASIL"

# Exact product name (accent-stripped, upper) -> our product name.
PRODUCTS = {
    "GASOLINA COMUM": "petrol",
    "GASOLINA ADITIVADA": "petrol_premium",
    "OLEO DIESEL": "diesel",
    "OLEO DIESEL S10": "diesel_s10",
    "ETANOL HIDRATADO": "ethanol",
}

# Only rows carrying this unit are per-litre prices.
LITRE_UNITS = ("R$/L", "RS/L")

MIN_PRICE, MAX_PRICE = 0.5, 100.0


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _plain(value) -> str:
    """Upper-case, accent-stripped text for robust comparison."""
    text = str(value or "").strip()
    stripped = unicodedata.normalize("NFKD", text)
    return "".join(c for c in stripped if not unicodedata.combining(c)).upper()


def _workbook_link(page_html: str) -> str | None:
    soup = BeautifulSoup(page_html, "html.parser")
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        if LINK_MARKER in href.lower() and href.lower().endswith(".xlsx"):
            return ANP_BASE + href if href.startswith("/") else href
    return None


def _header_row(rows) -> int | None:
    for index, row in enumerate(rows):
        cells = [_plain(c) for c in row]
        if "PRODUTO" in cells and any(c.startswith("UNIDADE") for c in cells):
            return index
    return None


def _columns(header) -> dict:
    """Locate the product, unit and average-price columns by header text."""
    found = {}
    for index, cell in enumerate(header):
        name = _plain(cell)
        if name == "PRODUTO":
            found["product"] = index
        elif name.startswith("UNIDADE"):
            found["unit"] = index
        elif name.startswith("PRECO MEDIO"):
            found["price"] = index
    return found


def _parse_workbook(data: bytes) -> list[LocalPrice]:
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    if NATIONAL_SHEET not in workbook.sheetnames:
        raise ValueError("ANP workbook has no " + NATIONAL_SHEET + " sheet")
    rows = list(workbook[NATIONAL_SHEET].iter_rows(values_only=True))

    header_index = _header_row(rows)
    if header_index is None:
        raise ValueError("ANP sheet has no recognisable header row")
    columns = _columns(rows[header_index])
    if not {"product", "unit", "price"} <= set(columns):
        raise ValueError("ANP sheet is missing product, unit or price columns")

    fetched = _now_utc()
    prices = []
    for row in rows[header_index + 1:]:
        if columns["price"] >= len(row):
            continue
        if _plain(row[columns["unit"]]) not in LITRE_UNITS:
            continue
        product = PRODUCTS.get(_plain(row[columns["product"]]))
        if product is None:
            continue
        try:
            value = float(row[columns["price"]])
        except (TypeError, ValueError):
            continue
        if not MIN_PRICE <= value <= MAX_PRICE:
            continue
        prices.append(LocalPrice(
            country_code="BR", product=product, price=round(value, 6),
            currency="BRL", fetched_utc=fetched, source="gov.br/anp",
        ))
    if not prices:
        raise ValueError("ANP workbook produced no usable prices")
    return prices


def fetch() -> list[LocalPrice]:
    """Return the latest Brazilian national pump prices, BRL per litre."""
    link = _workbook_link(http.get(ANP_PAGE_URL).text)
    if link is None:
        raise ValueError("No weekly summary workbook found on " + ANP_PAGE_URL)
    return _parse_workbook(http.get(link).content)
