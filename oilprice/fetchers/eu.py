"""European Union retail fuel prices (EUR per litre), all member states.

The European Commission's Weekly Oil Bulletin publishes one spreadsheet
covering every member state, which is why a single fetcher here yields
rows for ~27 countries. The pipeline keys scrapers by country only for
logging; each LocalPrice carries its own country_code, so a multi-country
source needs no special handling.

Two things about the source shape the parsing:

  * Prices are quoted per 1000 litres, except heavy fuel oil which is per
    tonne. Only the "1000 l" columns are read, so a tonne price can never
    be stored as if it were a litre price.
  * Every figure is converted to EUR by the Commission, including for
    member states outside the euro (Poland, Sweden, Denmark, Hungary,
    Czechia, Bulgaria, Romania). Rows are therefore labelled EUR, not the
    national currency.

The download URL carries a document UUID, so the link is discovered from
the bulletin page rather than hardcoded.
"""

import io
import logging
from datetime import datetime, timezone

import openpyxl
from bs4 import BeautifulSoup

from ..countries import load_countries
from ..models import LocalPrice
from . import http

log = logging.getLogger(__name__)

EU_PAGE_URL = "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
EU_BASE = "https://energy.ec.europa.eu"

# We want the consumer-facing file: prices including taxes and duties.
LINK_KEYWORDS = ("with taxes", "with%20taxes")

# Column header keyword -> our product name.
COLUMN_PRODUCTS = (
    ("euro-super", "petrol"),
    ("gas oil automobile", "diesel"),
    ("gas oil de chauffage", "heating_oil"),
    ("gpl", "lpg"),
    ("lpg", "lpg"),
)

# Only columns quoted in this unit are read; heavy fuel oil uses tonnes.
LITRE_UNIT = "1000 l"
LITRES_PER_UNIT = 1000.0

# Sanity bounds for EUR per 1000 litres.
MIN_PRICE, MAX_PRICE = 100.0, 10000.0


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _xlsx_link(page_html: str) -> str | None:
    soup = BeautifulSoup(page_html, "html.parser")
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        lowered = href.lower()
        if ".xlsx" not in lowered:
            continue
        if any(k in lowered for k in LINK_KEYWORDS):
            return EU_BASE + href if href.startswith("/") else href
    return None


def _product_for(header: str) -> str | None:
    lowered = str(header or "").strip().lower()
    for keyword, product in COLUMN_PRODUCTS:
        if keyword in lowered:
            return product
    return None


def _price_columns(header_row, unit_row) -> dict[int, str]:
    """Map column index -> product, for per-litre columns only."""
    columns: dict[int, str] = {}
    for index in range(1, len(header_row)):
        unit = str(unit_row[index] or "").strip().lower() if index < len(unit_row) else ""
        if unit != LITRE_UNIT:
            continue
        product = _product_for(header_row[index])
        if product is not None:
            columns[index] = product
    return columns


def _parse_workbook(data: bytes) -> list[LocalPrice]:
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 3:
        raise ValueError("EU bulletin sheet has no data rows")

    columns = _price_columns(rows[0], rows[1])
    if not columns:
        raise ValueError("EU bulletin sheet has no per-litre price columns")

    by_name = {info["name"]: code for code, info in load_countries().items()}
    fetched = _now_utc()
    prices: list[LocalPrice] = []
    unresolved = []

    for row in rows[2:]:
        name = str(row[0] or "").strip()
        if not name:
            continue
        code = by_name.get(name)
        if code is None:
            # Aggregate rows (EU27, Euro Area) land here and are skipped.
            unresolved.append(name)
            continue
        for index, product in columns.items():
            if index >= len(row) or row[index] in (None, ""):
                continue
            try:
                per_unit = float(row[index])
            except (TypeError, ValueError):
                continue
            if not MIN_PRICE <= per_unit <= MAX_PRICE:
                continue
            prices.append(LocalPrice(
                country_code=code, product=product,
                price=round(per_unit / LITRES_PER_UNIT, 6),
                currency="EUR", fetched_utc=fetched,
                source="energy.ec.europa.eu",
            ))
    if unresolved:
        log.debug("EU bulletin rows skipped (not countries): %s", unresolved)
    if not prices:
        raise ValueError("EU bulletin produced no usable prices")
    return prices


def fetch() -> list[LocalPrice]:
    """Return latest pump prices for every EU member state, EUR per litre."""
    link = _xlsx_link(http.get(EU_PAGE_URL).text)
    if link is None:
        raise ValueError("No 'with taxes' XLSX link found on " + EU_PAGE_URL)
    return _parse_workbook(http.get(link).content)
