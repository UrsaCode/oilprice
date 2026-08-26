"""What the sellers charge: the OPEC basket, and Dubai.

BRENT IS NOT WHAT SOUTH ASIA BUYS. Brent is North Sea crude and WTI is
American, and between them they are the two benchmarks everybody quotes — but
a refinery in Karachi or Chennai is buying Arab Light from Saudi Arabia and
Murban from the UAE, and those cargoes are priced off **Dubai**, not off
Brent. A page that says "the crude in your litre costs this much" using Brent
alone is answering with the wrong seller's price, and the two are not
interchangeable: the OPEC basket has traded five dollars a barrel clear of
Brent on the day this module was written.

TWO BENCHMARKS, AND WHAT EACH ONE IS.

``OPEC`` is the OPEC Reference Basket, the daily average of the twelve member
crudes — Arab Light for Saudi Arabia, Murban for the UAE, Iran Heavy, Basra
Medium, Kuwait Export and the rest. It is one blended figure and not a
per-country price, which matters: it is the closest free daily reading of
what the exporting side is charging, and it is *not* Saudi Arabia's price.

``DUBAI`` is the Dubai benchmark as the World Bank publishes it, and it is the
one Asian cargoes are actually priced against. It is a **monthly average**,
so it is the same number all month and the ``source`` says which month it is
for. A monthly figure repeated across sixty runs is not sixty readings.

WHY THESE ARE NOT FALLBACKS FOR EACH OTHER, unlike the chain in
``international.py``. Yahoo, FRED and the EIA republication are three ways of
reading the same Brent, so the first that answers wins and the rest are never
asked. These are different quantities from different publishers. Each is
fetched on its own and one being down costs only itself; the run is marked
partial and the other still stores.

PER-SELLER PRICES ARE NOT FREELY PUBLISHED, and this module does not pretend
otherwise. Saudi Aramco announces its Official Selling Prices monthly as a
press release with no feed behind it, ADNOC's Murban settles on ICE Futures
Abu Dhabi whose data pages refuse anonymous clients, and Iranian cargoes are
not openly priced at all. Somebody who needs Arab Light on its own needs a
paid assessment; what is here is the basket those prices go into.
"""

import io
import logging
import re
from datetime import datetime, timezone

import openpyxl
from defusedxml import ElementTree

from ..models import BenchmarkQuote
from . import http

log = logging.getLogger(__name__)

# Daily, back to 2003, as XML. Served from behind a filter that reads the TLS
# handshake, which is why this one source uses http.browser_get.
OPEC_URL = "https://www.opec.org/basket/basketDayArchives.xml"

# The Pink Sheet. The file's URL carries a content hash that changes whenever
# it is republished, so it is discovered from the landing page and the last
# known address is only the fallback.
WORLD_BANK_PAGE = "https://www.worldbank.org/en/research/commodity-markets"
WORLD_BANK_FILE = (
    "https://thedocs.worldbank.org/en/doc/"
    "5d903e848db1d1b83e0ec8f744e55570-0350012021/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)
WORLD_BANK_PATTERN = re.compile(
    r"https://thedocs\.worldbank\.org[^\"'\s]*CMO-Historical-Data-Monthly\.xlsx"
)
WORLD_BANK_SHEET = "Monthly Prices"
DUBAI_COLUMN = "Crude oil, Dubai"

# USD per barrel a crude benchmark can plausibly be. Brent has traded under
# twenty and over a hundred and forty inside this record's lifetime, so the
# bounds are wide on purpose: they exist to reject a parse that read a year
# or an index, not to have an opinion about the market.
MIN_USD, MAX_USD = 5.0, 400.0


def _now_utc() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _plausible(value: float) -> bool:
    return MIN_USD <= value <= MAX_USD


# --- the OPEC Reference Basket ------------------------------------------

def _from_opec() -> BenchmarkQuote:
    """The newest daily basket price OPEC has published.

    Entries are ``<BasketList data="2026-08-25" val="90.28"/>`` under a
    namespace, so they are matched by local name rather than by the namespace
    URI — that URI is a tempuri.org placeholder and pinning it would make the
    parser fail the day somebody tidies it up.

    Parsed with defusedxml rather than the standard library: this is remote
    XML from a host outside our control, and the stdlib parser's defences
    against entity-expansion attacks are not something to have to reason about
    per call site.
    """
    body = http.browser_get(OPEC_URL).content
    root = ElementTree.fromstring(body)

    newest_date, newest_value = None, None
    for node in root.iter():
        if not node.tag.endswith("BasketList"):
            continue
        when = node.get("data")
        raw = node.get("val")
        if not when or not raw:
            continue
        try:
            value = float(raw)
        except ValueError:
            continue
        if not _plausible(value):
            continue
        if newest_date is None or when > newest_date:
            newest_date, newest_value = when, value

    if newest_date is None:
        raise ValueError("OPEC basket feed carried no usable entry")

    return BenchmarkQuote(
        benchmark="OPEC",
        price_usd=newest_value,
        fetched_utc=_now_utc(),
        source=f"opec.org (reference basket, {newest_date})",
    )


# --- Dubai, from the World Bank Pink Sheet -------------------------------

def _world_bank_file(page_html: str) -> str:
    """The monthly workbook the landing page currently links, or the last one."""
    found = WORLD_BANK_PATTERN.search(page_html or "")
    return found.group(0) if found else WORLD_BANK_FILE


def _newest_dubai(data: bytes) -> tuple[str, float]:
    """The last month the sheet prices Dubai for, and the price.

    The header sits several rows down under a title block, so the column is
    found by its published name rather than by position: the World Bank adds
    commodities, and a fixed index would quietly start reading coal.
    """
    book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if WORLD_BANK_SHEET not in book.sheetnames:
        raise ValueError(f"no '{WORLD_BANK_SHEET}' sheet in the Pink Sheet")
    sheet = book[WORLD_BANK_SHEET]

    column = None
    newest_month, newest_value = None, None
    for row in sheet.iter_rows(values_only=True):
        if column is None:
            for index, cell in enumerate(row):
                if isinstance(cell, str) and cell.strip() == DUBAI_COLUMN:
                    column = index
                    break
            continue

        month = row[0]
        if not isinstance(month, str) or "M" not in month:
            continue          # the units row, and any blank line
        if column >= len(row):
            continue
        try:
            value = float(row[column])
        except (TypeError, ValueError):
            continue          # ".." is how the sheet writes a month it has no price for
        if not _plausible(value):
            continue
        if newest_month is None or month > newest_month:
            newest_month, newest_value = month, value

    if newest_month is None:
        raise ValueError(f"'{DUBAI_COLUMN}' had no usable month")
    return newest_month, newest_value


def _from_world_bank() -> BenchmarkQuote:
    page = ""
    try:
        page = http.get(WORLD_BANK_PAGE).text
    except Exception as exc:
        log.warning("World Bank landing page unreachable (%s), using last URL", exc)

    url = _world_bank_file(page)
    month, value = _newest_dubai(http.get(url).content)

    return BenchmarkQuote(
        benchmark="DUBAI",
        price_usd=value,
        fetched_utc=_now_utc(),
        # The month is in the source string because a monthly average stored
        # twice a day is one reading repeated, and a reader has to be able to
        # tell that from a series that moved.
        source=f"worldbank.org (Dubai, {month} monthly average)",
    )


def fetch() -> list[BenchmarkQuote]:
    """Both seller-side benchmarks, each independently best-effort."""
    quotes = []
    for name, read in (("OPEC", _from_opec), ("DUBAI", _from_world_bank)):
        try:
            quotes.append(read())
        except Exception as exc:
            log.warning("%s benchmark fetch failed: %s", name, exc)
    if not quotes:
        raise RuntimeError("No seller-side benchmark could be fetched")
    return quotes
