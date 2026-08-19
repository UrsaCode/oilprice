"""Offline tests: full pipeline with mocked network, plus parser checks.

Run with:  python -m pytest tests/ -v
"""

import json
from datetime import datetime, timezone

import pytest

from oilprice import config, db, pipeline
from oilprice.fetchers import pakistan
from oilprice.models import BenchmarkQuote, LocalPrice

NOW = datetime.now(timezone.utc).isoformat(timespec="seconds")

FAKE_QUOTES = [
    BenchmarkQuote("BRENT", 80.5, NOW, "test"),
    BenchmarkQuote("WTI", 76.2, NOW, "test"),
]
FAKE_RATES = {"PKR": 280.0, "USD": 1.0, "EUR": 0.92}
FAKE_LOCAL = [
    LocalPrice("PK", "petrol", 275.6, "PKR", NOW, "test"),
    LocalPrice("PK", "diesel", 285.3, "PKR", NOW, "test"),
]


@pytest.fixture
def isolated_data(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "oilprice.db")
    monkeypatch.setattr(config, "CSV_DIR", tmp_path / "csv")
    monkeypatch.setattr(config, "SNAPSHOT_DIR", tmp_path / "snapshots")
    return tmp_path


def test_full_run_with_mocked_sources(isolated_data, monkeypatch):
    monkeypatch.setattr(pipeline.international, "fetch", lambda: FAKE_QUOTES)
    monkeypatch.setattr(pipeline.fx, "fetch", lambda: (FAKE_RATES, "test-fx"))
    monkeypatch.setattr(pipeline, "LOCAL_SCRAPERS", {"PK": lambda: FAKE_LOCAL})

    assert pipeline.run() == "ok"

    conn = db.connect()
    intl = conn.execute("SELECT * FROM international_prices").fetchall()
    assert {r["benchmark"] for r in intl} == {"BRENT", "WTI"}

    pk = conn.execute(
        "SELECT * FROM benchmark_local WHERE country_code = 'PK' "
        "AND benchmark = 'BRENT'"
    ).fetchone()
    assert pk["currency"] == "PKR"
    assert pk["price_per_barrel"] == pytest.approx(80.5 * 280.0)
    assert pk["price_per_litre"] == pytest.approx(
        80.5 * 280.0 / config.LITRES_PER_BARREL
    )

    pump = conn.execute(
        "SELECT price FROM local_prices WHERE country_code='PK' "
        "AND product='petrol'"
    ).fetchone()
    assert pump["price"] == 275.6

    # Second run in the same slot is skipped, not duplicated.
    assert pipeline.run() == "skipped"
    assert conn.execute(
        "SELECT COUNT(*) c FROM international_prices"
    ).fetchone()["c"] == 2

    # Exports exist.
    assert (isolated_data / "csv" / "international.csv").exists()
    snapshots = list((isolated_data / "snapshots").glob("*.json"))
    assert len(snapshots) == 1
    snap = json.loads(snapshots[0].read_text())
    assert snap["international_usd_per_barrel"]["BRENT"] == 80.5


def _boom():
    raise RuntimeError("source down")


def test_partial_when_local_scrape_fails(isolated_data, monkeypatch):
    monkeypatch.setattr(pipeline.international, "fetch", lambda: FAKE_QUOTES)
    monkeypatch.setattr(pipeline.fx, "fetch", lambda: (FAKE_RATES, "test-fx"))
    monkeypatch.setattr(pipeline, "LOCAL_SCRAPERS", {"PK": _boom})

    assert pipeline.run() == "partial"


def test_international_failure_still_saves_fx_and_local(isolated_data,
                                                        monkeypatch):
    monkeypatch.setattr(pipeline.international, "fetch", _boom)
    monkeypatch.setattr(pipeline.fx, "fetch", lambda: (FAKE_RATES, "test-fx"))
    monkeypatch.setattr(pipeline, "LOCAL_SCRAPERS", {"PK": lambda: FAKE_LOCAL})

    assert pipeline.run() == "partial"

    conn = db.connect()
    assert conn.execute("SELECT COUNT(*) c FROM local_prices").fetchone()["c"] == 2
    assert conn.execute("SELECT COUNT(*) c FROM fx_rates").fetchone()["c"] == 3
    # No benchmark conversion possible without international quotes.
    assert conn.execute(
        "SELECT COUNT(*) c FROM benchmark_local"
    ).fetchone()["c"] == 0


def test_failed_when_every_source_fails(isolated_data, monkeypatch):
    monkeypatch.setattr(pipeline.international, "fetch", _boom)
    monkeypatch.setattr(pipeline.fx, "fetch", _boom)
    monkeypatch.setattr(pipeline, "LOCAL_SCRAPERS", {"PK": _boom})

    assert pipeline.run() == "failed"
    # Nothing collected -> no exports written.
    assert not (isolated_data / "snapshots").exists()


def test_fred_parser(monkeypatch):
    from oilprice.fetchers import international

    csv_by_url = {
        international.FRED_URL.format(series="DCOILBRENTEU"):
            "DATE,DCOILBRENTEU\n2026-08-11,79.10\n2026-08-12,80.25\n2026-08-13,.\n",
        international.FRED_URL.format(series="DCOILWTICO"):
            "DATE,DCOILWTICO\n2026-08-12,76.40\n",
    }

    class FakeResp:
        def __init__(self, text):
            self.text = text

    monkeypatch.setattr(
        international.http, "get", lambda url: FakeResp(csv_by_url[url])
    )
    quotes = {q.benchmark: q for q in international._from_fred()}
    # "." (missing day) rows are skipped; latest numeric value wins.
    assert quotes["BRENT"].price_usd == 80.25
    assert quotes["WTI"].price_usd == 76.40


def test_github_dataset_parser(monkeypatch):
    from oilprice.fetchers import international

    csv_by_url = {
        international.GITHUB_DATASET_URL.format(
            branch="master", file="brent-daily.csv"):
            "Date,Price\n2026-08-10,92.74\n2026-08-11,93.26\n",
        international.GITHUB_DATASET_URL.format(
            branch="master", file="wti-daily.csv"):
            "Date,Price\n2026-08-11,84.77\n",
    }

    class FakeResp:
        def __init__(self, text):
            self.text = text

    def fake_get(url):
        if url not in csv_by_url:
            raise RuntimeError("404")
        return FakeResp(csv_by_url[url])

    monkeypatch.setattr(international.http, "get", fake_get)
    quotes = {q.benchmark: q for q in international._from_github_dataset()}
    assert quotes["BRENT"].price_usd == 93.26
    assert quotes["WTI"].price_usd == 84.77


def test_pakistan_table_parser(monkeypatch):
    html = """
    <table>
      <tr><th>Product</th><th>Price</th></tr>
      <tr><td>Premier Euro 5</td><td>Rs. 275.60 /Ltr</td></tr>
      <tr><td>Hi-Cetane Diesel</td><td>Rs. 285.34 /Ltr</td></tr>
      <tr><td>Kerosene Oil</td><td>183.16</td></tr>
    </table>
    """

    class FakeResp:
        text = html

    monkeypatch.setattr(pakistan.http, "get", lambda url: FakeResp())
    prices = {p.product: p for p in pakistan._scrape_tables("http://x", "test")}
    assert prices["petrol"].price == 275.60
    assert prices["diesel"].price == 285.34
    assert prices["kerosene"].price == 183.16
    assert all(p.currency == "PKR" for p in prices.values())


def test_pakistan_pso_current_markup(monkeypatch):
    """Row-oriented table as served by psopk.com/en/fuels/fuel-prices."""
    html = """
    <table>
      <tr><th>Product Name</th><th>Rs./Litre</th></tr>
      <tr><td>PREMIER EURO 5</td><td>Rs.325.43/Ltr</td></tr>
      <tr><td>HI-CETANE DIESEL EURO 5</td><td>Rs.383.95/Ltr</td></tr>
      <tr><td>LDO</td><td>Rs.249.03/Ltr</td></tr>
      <tr><td>SKO</td><td>Rs.289.75/Ltr</td></tr>
      <tr><td>JP-1</td><td>Rs.324.8/Ltr</td></tr>
    </table>
    """

    class FakeResp:
        text = html

    monkeypatch.setattr(pakistan.http, "get", lambda url: FakeResp())
    prices = {p.product: p for p in pakistan._scrape_tables("http://x", "test")}
    assert prices["petrol"].price == 325.43
    assert prices["diesel"].price == 383.95
    assert prices["light_diesel"].price == 249.03
    assert prices["kerosene"].price == 289.75
    # Jet fuel is not one of our products.
    assert set(prices) == {"petrol", "diesel", "light_diesel", "kerosene"}


def test_pakistan_column_oriented_table(monkeypatch):
    """hamariweb lists products as headers with one row per date."""
    html = """
    <table>
      <tr><th>Date</th><th>Petrol</th><th>HS Diesel</th>
          <th>Light Diesel</th><th>Kerosene Oil</th></tr>
      <tr><td>Aug 14, 2026</td><td>325.43</td><td>383.95</td>
          <td>249.03</td><td>289.75</td></tr>
      <tr><td>Aug 13, 2026</td><td>324.98</td><td>382.79</td>
          <td>247.82</td><td>289.27</td></tr>
    </table>
    """

    class FakeResp:
        text = html

    monkeypatch.setattr(pakistan.http, "get", lambda url: FakeResp())
    prices = {p.product: p for p in pakistan._scrape_tables("http://x", "test")}
    # Most recent row (first) wins; the date column is never read as a price.
    assert prices["petrol"].price == 325.43
    assert prices["diesel"].price == 383.95
    assert prices["light_diesel"].price == 249.03
    assert prices["kerosene"].price == 289.75


def test_light_diesel_not_classified_as_diesel():
    """'Light Diesel' contains 'diesel'; the more specific product must win."""
    assert pakistan._classify("Light Diesel") == "light_diesel"
    assert pakistan._classify("Light Diesel Oil") == "light_diesel"
    assert pakistan._classify("LDO") == "light_diesel"
    assert pakistan._classify("HS Diesel") == "diesel"
    assert pakistan._classify("Hi-Cetane Diesel Euro 5") == "diesel"
    assert pakistan._classify("Premier Euro 5") == "petrol"
    assert pakistan._classify("Kerosene Oil") == "kerosene"
    assert pakistan._classify("Date") is None
    assert pakistan._classify("JP-1") is None


# --- United States (EIA) -----------------------------------------------

EIA_HTML = """
<table>
  <caption>U.S. Regular Gasoline Prices*(dollars per gallon) full history XLS</caption>
  <tr><td></td><td></td><td></td><td></td><td>Change from</td></tr>
  <tr><td></td><td>07/27/26</td><td>08/03/26</td><td>08/10/26</td>
      <td>2 year ago</td><td>year ago</td><td>week ago</td></tr>
  <tr><td>U.S.</td><td>4.096</td><td>4.079</td><td>4.006</td>
      <td>0.592</td><td>0.888</td><td>-0.073</td></tr>
  <tr><td>East Coast (PADD1)</td><td>3.997</td><td>3.944</td><td>3.884</td>
      <td>0.558</td><td>0.879</td><td>-0.060</td></tr>
</table>
<table>
  <caption>States</caption>
  <tr><td></td><td>07/27/26</td><td>08/03/26</td><td>08/10/26</td></tr>
  <tr><td>California</td><td>5.489</td><td>5.495</td><td>5.428</td></tr>
</table>
<table>
  <caption>U.S. On-Highway Diesel Fuel Prices*(dollars per gallon) full history XLS</caption>
  <tr><td></td><td></td><td></td><td></td><td>Change from</td></tr>
  <tr><td></td><td>07/27/26</td><td>08/03/26</td><td>08/10/26</td>
      <td>2 year ago</td><td>year ago</td><td>week ago</td></tr>
  <tr><td>U.S.</td><td>5.313</td><td>5.348</td><td>5.257</td>
      <td>1.553</td><td>1.503</td><td>-0.091</td></tr>
</table>
"""


def test_usa_parser(monkeypatch):
    from oilprice.fetchers import usa

    class FakeResp:
        text = EIA_HTML

    monkeypatch.setattr(usa.http, "get", lambda url: FakeResp())
    prices = {p.product: p for p in usa.fetch()}

    # Latest date column (08/10/26) wins, converted from USD/gallon to /litre.
    assert prices["petrol"].price == pytest.approx(
        4.006 / config.LITRES_PER_US_GALLON, abs=1e-6)
    assert prices["diesel"].price == pytest.approx(
        5.257 / config.LITRES_PER_US_GALLON, abs=1e-6)
    assert all(p.currency == "USD" and p.unit == "litre" and p.country_code == "US"
               for p in prices.values())
    assert set(prices) == {"petrol", "diesel"}


def test_usa_ignores_change_from_columns(monkeypatch):
    """The 'Change from' deltas sit after the dates and must never be read."""
    from oilprice.fetchers import usa

    class FakeResp:
        text = EIA_HTML

    monkeypatch.setattr(usa.http, "get", lambda url: FakeResp())
    prices = {p.product: p for p in usa.fetch()}
    # 0.888 (a year-ago delta) would convert to ~0.235; the real value is ~1.06.
    assert prices["petrol"].price > 1.0


# --- United Kingdom (DESNZ) --------------------------------------------

UK_CSV = chr(0xFEFF) + chr(10).join([
    "Date,ULSP (Ultra low sulphur unleaded petrol) Pump price in pence/litre,"
    "ULSD (Ultra low sulphur diesel) Pump price in pence/litre,"
    "ULSP duty,ULSD duty,ULSP VAT,ULSD VAT",
    "27/07/2026,156.13,173.97,52.95,52.95,20,20",
    "03/08/2026,159.89,179.19,52.95,52.95,20,20",
    "10/08/2026,162.15,181.97,52.95,52.95,20,20",
]) + chr(10)

UK_PAGE_HTML = (
    '<html><body><a href="/media/abc123/Weekly_Fuel_Prices.xlsx">XLSX</a>'
    '<a href="/media/def456/CSV__2018_-__.csv">CSV</a></body></html>'
)


def _uk_fake_get(csv_text):
    """Serve the statistics page, then the CSV it links to."""
    from oilprice.fetchers import uk as _uk

    def fake_get(url):
        class Resp:
            text = UK_PAGE_HTML if url == _uk.UK_PAGE_URL else csv_text
        return Resp()
    return fake_get


def test_uk_parser(monkeypatch):
    from oilprice.fetchers import uk

    monkeypatch.setattr(uk.http, "get", _uk_fake_get(UK_CSV))
    prices = {p.product: p for p in uk.fetch()}

    # Last row is the most recent; pence converted to pounds.
    assert prices["petrol"].price == pytest.approx(1.6215)
    assert prices["diesel"].price == pytest.approx(1.8197)
    assert all(p.currency == "GBP" and p.country_code == "GB"
               for p in prices.values())


def test_uk_skips_trailing_blank_rows(monkeypatch):
    from oilprice.fetchers import uk

    monkeypatch.setattr(
        uk.http, "get", _uk_fake_get(UK_CSV + ",,,,,," + chr(10) + chr(10)))
    prices = {p.product: p for p in uk.fetch()}
    assert prices["petrol"].price == pytest.approx(1.6215)


# --- European Union (Weekly Oil Bulletin) ------------------------------

def _eu_workbook():
    """Build a workbook shaped like the real Weekly Oil Bulletin."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["in EUR", "Euro-super 95  (I)", "Gas oil automobile Aut",
               " Gas oil de chauffage ", " Fuel oil - Schweres H",
               " Fuel oil -Schweres He", "GPL pour moteur LPG mo"])
    ws.append(["2026-08-10 00:00:00", "1000 l", "1000 l", "1000 l", "t", "t", "1000 l"])
    ws.append(["Germany", 2143, 2149, 1315.7, None, None, 1090.38])
    ws.append(["France", 2014.06, 2169.04, 1646.17, None, None, 1018.11])
    ws.append(["Poland", 1696.19, 1863.79, 1390.73, 697.67, 527.21, 700.99])
    ws.append(["Austria", 1721, 1965, 1542.87])            # short row
    ws.append(["Malta", 1340, 1210, 1000])
    ws.append(["CE/EC/EG EUR27_2020 (I", 1910.68, 2012.67, 1416.59])   # aggregate
    ws.append(["CE/EC/EG Euro Area 20 ", 1973.22, 2049.22, 1399.60])   # aggregate
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


EU_PAGE_HTML = (
    '<html><body>'
    '<a href="/document/download/uuid-1_en?filename=Oil_Bulletin_Duties.xlsx">duties</a>'
    '<a href="/document/download/uuid-2_en?filename=Weekly%20prices%20with%20taxes.xlsx">'
    'prices with taxes</a>'
    '</body></html>'
)


def _eu_fake_get():
    """Serve the bulletin page, then the workbook it links to."""
    from oilprice.fetchers import eu as _eu
    book = _eu_workbook()

    def fake_get(url):
        class Resp:
            text = EU_PAGE_HTML
            content = book
        return Resp()
    return fake_get


def test_eu_parser(monkeypatch):
    from oilprice.fetchers import eu

    monkeypatch.setattr(eu.http, "get", _eu_fake_get())
    prices = eu.fetch()
    by_country = {}
    for p in prices:
        by_country.setdefault(p.country_code, {})[p.product] = p.price

    # EUR per 1000 litres -> EUR per litre.
    assert by_country["DE"]["petrol"] == pytest.approx(2.143)
    assert by_country["DE"]["diesel"] == pytest.approx(2.149)
    assert by_country["FR"]["petrol"] == pytest.approx(2.01406)
    assert by_country["MT"]["petrol"] == pytest.approx(1.340)

    # Country names resolve to ISO codes via the shared reference data.
    assert "PL" in by_country and "AT" in by_country

    # Aggregate rows (EU27, Euro Area) are not countries.
    assert not [p for p in prices if p.country_code.startswith("CE")]
    assert len(by_country) == 5


def test_eu_prices_are_labelled_eur_not_national_currency(monkeypatch):
    """The bulletin converts everything to EUR, including for non-euro states."""
    from oilprice.fetchers import eu

    monkeypatch.setattr(eu.http, "get", _eu_fake_get())
    poland = [p for p in eu.fetch() if p.country_code == "PL"]
    assert poland, "Poland should be present"
    # Poland's currency is PLN, but these figures are euro-denominated.
    assert all(p.currency == "EUR" for p in poland)


def test_eu_skips_tonne_priced_heavy_fuel_oil(monkeypatch):
    """Heavy fuel oil is quoted per tonne; mixing units would corrupt the store."""
    from oilprice.fetchers import eu

    monkeypatch.setattr(eu.http, "get", _eu_fake_get())
    prices = eu.fetch()
    assert all(p.unit == "litre" for p in prices)
    assert all("fuel_oil" not in p.product for p in prices)
    # Poland has heavy-fuel-oil values in the sheet; they must not appear.
    pl = {p.product for p in prices if p.country_code == "PL"}
    assert pl == {"petrol", "diesel", "heating_oil", "lpg"}


def test_all_scrapers_registered():
    from oilprice.fetchers import LOCAL_SCRAPERS
    assert set(LOCAL_SCRAPERS) == {"PK", "US", "GB", "EU", "BR", "EG",
                                   "MX", "BD", "IN"}


def test_multi_country_scraper_logs_every_country(isolated_data, monkeypatch,
                                                  caplog):
    """A multi-country source must not be reported as one country alone."""
    multi = [
        LocalPrice("DE", "petrol", 2.14, "EUR", NOW, "test"),
        LocalPrice("FR", "petrol", 2.01, "EUR", NOW, "test"),
        LocalPrice("PL", "petrol", 1.69, "EUR", NOW, "test"),
    ]
    monkeypatch.setattr(pipeline.international, "fetch", lambda: FAKE_QUOTES)
    monkeypatch.setattr(pipeline.fx, "fetch", lambda: (FAKE_RATES, "test-fx"))
    monkeypatch.setattr(pipeline, "LOCAL_SCRAPERS", {"EU": lambda: multi})

    with caplog.at_level("INFO"):
        assert pipeline.run() == "ok"
    assert "3 countries" in caplog.text

    conn = db.connect()
    stored = conn.execute(
        "SELECT DISTINCT country_code FROM local_prices").fetchall()
    assert {r["country_code"] for r in stored} == {"DE", "FR", "PL"}


# --- Brazil (ANP) ------------------------------------------------------

def _br_workbook():
    """Workbook shaped like the ANP weekly resumo, BRASIL sheet."""
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("CAPITAIS")
    ws.append(["ignored"])
    ws = wb.create_sheet("BRASIL")
    for _ in range(9):
        ws.append([])
    ws.append(["DATA INICIAL", "DATA FINAL", "BRASIL", "PRODUTO",
               "NÚMERO DE POSTOS PESQUISADOS", "UNIDADE DE MEDIDA",
               "PREÇO MÉDIO REVENDA", "DESVIO PADRÃO REVENDA"])
    rows = [
        ("ETANOL HIDRATADO", "R$/l", 3.91),
        ("GASOLINA ADITIVADA", "R$/l", 6.73),
        ("GASOLINA COMUM", "R$/l", 6.53),
        ("GLP", "R$/13kg", 113.89),
        ("GNV", "R$/m3", 4.67),
        ("OLEO DIESEL", "R$/l", 6.51),
        ("OLEO DIESEL S10", "R$/l", 6.91),
    ]
    for product, unit, price in rows:
        ws.append(["2026-08-09", "2026-08-15", "BRASIL", product, 100,
                   unit, price, 0.4])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


BR_PAGE_HTML = (
    '<html><body>'
    '<a href="/anp/arquivos-lpc/2026/resumo_semanal_lpc_2026-08-09_2026-08-15.xlsx">resumo</a>'
    '<a href="/anp/arquivos-lpc/2026/revendas_lpc_2026-08-09_2026-08-15.xlsx">revendas</a>'
    '</body></html>'
)


def _br_fake_get():
    book = _br_workbook()

    def fake_get(url):
        class Resp:
            text = BR_PAGE_HTML
            content = book
        return Resp()
    return fake_get


def test_brazil_parser(monkeypatch):
    from oilprice.fetchers import brazil

    monkeypatch.setattr(brazil.http, "get", _br_fake_get())
    prices = {p.product: p for p in brazil.fetch()}

    assert prices["petrol"].price == pytest.approx(6.53)
    assert prices["diesel"].price == pytest.approx(6.51)
    assert prices["ethanol"].price == pytest.approx(3.91)
    assert all(p.currency == "BRL" and p.country_code == "BR" and p.unit == "litre"
               for p in prices.values())


def test_brazil_does_not_confuse_diesel_s10_with_diesel(monkeypatch):
    """'OLEO DIESEL S10' contains 'OLEO DIESEL' and must stay distinct."""
    from oilprice.fetchers import brazil

    monkeypatch.setattr(brazil.http, "get", _br_fake_get())
    prices = {p.product: p for p in brazil.fetch()}
    assert prices["diesel"].price == pytest.approx(6.51)
    assert prices["diesel_s10"].price == pytest.approx(6.91)


def test_brazil_skips_non_litre_units(monkeypatch):
    """GLP is priced per 13kg and GNV per cubic metre."""
    from oilprice.fetchers import brazil

    monkeypatch.setattr(brazil.http, "get", _br_fake_get())
    prices = brazil.fetch()
    assert all(p.unit == "litre" for p in prices)
    products = {p.product for p in prices}
    assert "lpg" not in products and "cng" not in products
    assert not [p for p in prices if p.price > 100]


# --- Egypt -------------------------------------------------------------

EG_HTML = """
<table>
  <tr><th></th><th>Type</th><th>Price</th><th>Unit</th></tr>
  <tr><td></td><td>Gasoline 80</td><td>20.75</td><td>piaster/liter</td></tr>
  <tr><td></td><td>Gasoline 92</td><td>22.25</td><td>LE/liter</td></tr>
  <tr><td></td><td>Gasoline 95</td><td>24</td><td>LE/liter</td></tr>
</table>
<table>
  <tr><th></th><th>Type</th><th>Price</th><th>Unit</th></tr>
  <tr><td></td><td>kerosene</td><td>20.5</td><td>LE/liter</td></tr>
  <tr><td></td><td>solar</td><td>20.5</td><td>LE/liter</td></tr>
  <tr><td></td><td>Gas stove</td><td>275</td><td>LE/cylinder</td></tr>
</table>
"""


def test_egypt_parser(monkeypatch):
    from oilprice.fetchers import egypt

    class FakeResp:
        text = EG_HTML

    monkeypatch.setattr(egypt.http, "get", lambda url: FakeResp())
    prices = {p.product: p for p in egypt.fetch()}

    assert prices["petrol_92"].price == pytest.approx(22.25)
    assert prices["petrol_95"].price == pytest.approx(24.0)
    assert prices["diesel"].price == pytest.approx(20.5)   # "solar"
    assert prices["kerosene"].price == pytest.approx(20.5)
    assert all(p.currency == "EGP" and p.country_code == "EG"
               for p in prices.values())


def test_egypt_excludes_per_cylinder_rows(monkeypatch):
    """Bottled gas is sold per cylinder, not per litre."""
    from oilprice.fetchers import egypt

    class FakeResp:
        text = EG_HTML

    monkeypatch.setattr(egypt.http, "get", lambda url: FakeResp())
    prices = egypt.fetch()
    assert all(p.unit == "litre" for p in prices)
    assert not [p for p in prices if p.price > 200]


# --- Mexico (CRE) ------------------------------------------------------

MX_XML = """<?xml version="1.0" encoding="utf-8"?>
<places>
  <place place_id="1">
    <gas_price type="regular">22.95</gas_price>
    <gas_price type="premium">27.90</gas_price>
    <gas_price type="diesel">25.00</gas_price>
  </place>
  <place place_id="2">
    <gas_price type="regular">24.50</gas_price>
    <gas_price type="premium">30.50</gas_price>
    <gas_price type="diesel">27.00</gas_price>
  </place>
  <place place_id="3">
    <gas_price type="regular">95.00</gas_price>
    <gas_price type="premium">28.00</gas_price>
  </place>
  <place place_id="4">
    <gas_price type="regular">999.00</gas_price>
  </place>
</places>
"""


def test_mexico_uses_median_not_mean(monkeypatch):
    """One absurd station reading must not drag the national figure."""
    from oilprice.fetchers import mexico

    class FakeResp:
        text = MX_XML
        content = MX_XML.encode("utf-8")

    monkeypatch.setattr(mexico.http, "get", lambda url: FakeResp())
    prices = {p.product: p for p in mexico.fetch()}

    # In-bounds readings 22.95 / 24.50 / 95.00 -> median 24.50; the mean
    # would be 47.48. The 999.00 station is rejected by the sanity bound
    # before the median is taken.
    assert prices["petrol"].price == pytest.approx(24.50)
    assert "3 stations" in prices["petrol"].source
    assert prices["diesel"].price == pytest.approx(26.00)
    assert all(p.currency == "MXN" and p.country_code == "MX"
               for p in prices.values())


def test_mexico_handles_byte_order_mark(monkeypatch):
    """The live feed begins with a BOM, which the XML parser rejects."""
    from oilprice.fetchers import mexico

    class FakeResp:
        content = (chr(0xFEFF) + MX_XML).encode("utf-8")

    monkeypatch.setattr(mexico.http, "get", lambda url: FakeResp())
    assert {p.product for p in mexico.fetch()} == {
        "petrol", "petrol_premium", "diesel"}


def test_mexico_labels_source_as_derived(monkeypatch):
    """The national figure is computed here, not published; say so."""
    from oilprice.fetchers import mexico

    class FakeResp:
        text = MX_XML
        content = MX_XML.encode("utf-8")

    monkeypatch.setattr(mexico.http, "get", lambda url: FakeResp())
    for p in mexico.fetch():
        assert "median" in p.source.lower()


# --- Bangladesh (BPC) --------------------------------------------------

BD_HTML = """
<table>
  <tr><th>নং</th><th>পণ্যের নাম</th>
      <th>স্থানীয় বিক্রয় মূল্য</th>
      <th>কার্যকরের তারিখ</th></tr>
  <tr><td>১</td><td>ডিজেল</td>
      <td>১১৫.০০ (টাকা/লিটার)</td><td>০১/০৬/২০২৬</td></tr>
  <tr><td>২</td><td>কেরোসিন</td>
      <td>১৩৫.০০ (টাকা/লিটার)</td><td>০১/০৬/২০২৬</td></tr>
  <tr><td>৩</td><td>অকটেন</td>
      <td>১৪৫.০০ (টাকা/লিটার)</td><td>০১/০৬/২০২৬</td></tr>
  <tr><td>৪</td><td>পেট্রোল</td>
      <td>১৪০.০০ (টাকা/লিটার)</td><td>০১/০৬/২০২৬</td></tr>
  <tr><td>৫</td><td>জেট এ-১</td>
      <td>১.০৩৫৮ (মা.ড/লিটার)</td><td>১০/০৮/২০২৬</td></tr>
  <tr><td>৬</td><td>এলপি গ্যাস</td>
      <td>৭৭৬.৫৩/সিলিন্ডার</td><td>২৩/০২/২০২৬</td></tr>
  <tr><td>৭</td><td>এলডিও</td>
      <td>১১৩.০০ (টাকা/লিটার)</td><td>০৭/০৭/২০২৬</td></tr>
</table>
"""


def test_bangladesh_parser(monkeypatch):
    from oilprice.fetchers import bangladesh

    class FakeResp:
        text = BD_HTML

    monkeypatch.setattr(bangladesh.http, "get",
                        lambda url, **kw: FakeResp())
    prices = {p.product: p for p in bangladesh.fetch()}

    # Bengali-Indic numerals are transliterated before parsing.
    assert prices["diesel"].price == pytest.approx(115.0)
    assert prices["kerosene"].price == pytest.approx(135.0)
    assert prices["petrol"].price == pytest.approx(140.0)
    assert prices["petrol_premium"].price == pytest.approx(145.0)
    assert prices["light_diesel"].price == pytest.approx(113.0)
    assert all(p.currency == "BDT" and p.country_code == "BD" and p.unit == "litre"
               for p in prices.values())


def test_bangladesh_rejects_dollar_priced_jet_fuel(monkeypatch):
    """Jet fuel is quoted per litre but in US dollars, not Taka."""
    from oilprice.fetchers import bangladesh

    class FakeResp:
        text = BD_HTML

    monkeypatch.setattr(bangladesh.http, "get",
                        lambda url, **kw: FakeResp())
    prices = bangladesh.fetch()
    # 1.0358 USD/litre must never be stored, in any currency.
    assert not [p for p in prices if p.price < 10]
    assert all(p.currency == "BDT" for p in prices)


def test_bangladesh_excludes_per_cylinder_rows(monkeypatch):
    from oilprice.fetchers import bangladesh

    class FakeResp:
        text = BD_HTML

    monkeypatch.setattr(bangladesh.http, "get",
                        lambda url, **kw: FakeResp())
    products = {p.product for p in bangladesh.fetch()}
    assert products == {"diesel", "kerosene", "petrol", "petrol_premium",
                        "light_diesel"}


def test_bangladesh_matches_both_bengali_encodings(monkeypatch):
    """ya-nukta may arrive precomposed or as base letter plus combining mark."""
    from oilprice.fetchers import bangladesh

    # Built from codepoints so the two spellings are genuinely different:
    # U+09DF, versus U+09AF followed by the combining nukta U+09BC.
    stem = "ফার্নেস অ"
    tail = "েল"
    precomposed = stem + chr(0x09DF) + tail
    decomposed = stem + chr(0x09AF) + chr(0x09BC) + tail
    assert precomposed != decomposed
    assert bangladesh._normalise(precomposed) == bangladesh._normalise(decomposed)

    for spelling in (precomposed, decomposed):
        html = ("<table><tr><td>১১</td><td>" + spelling
                + "</td><td>১০০.৩৯ (টাকা/লিটার)</td></tr></table>")

        class FakeResp:
            text = html

        monkeypatch.setattr(bangladesh.http, "get", lambda url, **kw: FakeResp())
        prices = {p.product: p for p in bangladesh.fetch()}
        assert prices["furnace_oil"].price == pytest.approx(100.39)


# --- India (PPAC) ------------------------------------------------------

IN_PAGE_HTML = (
    '<html><body>'
    '<a href="https://ppac.gov.in/download.php?file=whatsnew/'
    '1785476355_PP_9_a_DailyPriceMSHSD_Metro_31.07.2026.pdf">older</a>'
    '<a href="https://ppac.gov.in/uploads/page-images/'
    '1787118616_PP_9_a_DailyPriceMSHSD_Metro_19.08.2026.pdf">Current</a>'
    '</body></html>'
)

# Verbatim shape of the PDF text: petrol table left, diesel right.
IN_PDF_TEXT = chr(10).join([
    "Table Posted: 19-Aug-26",
    "Delhi Mumbai Chennai Kolkata Delhi Mumbai Chennai Kolkata",
    "19-Aug-26 102.12  111.21  107.77  113.51  19-Aug-26 95.20  97.83  99.55  99.82",
    "18-Aug-26 101.00  110.00  106.00  112.00  18-Aug-26 94.00  96.00  98.00  98.50",
    "Petroleum Planning & Analysis Cell",
    "Retail Selling Price of Petrol Retail Selling Price of Diesel",
    "Date of Revision Date of Revision",
    "(Rs./Litre) (Rs./Litre)",
])


def test_india_prefers_the_current_pdf_link():
    from oilprice.fetchers import india
    link = india._pdf_link(IN_PAGE_HTML)
    assert link.endswith("Metro_19.08.2026.pdf")


def test_india_reads_newest_row_and_delhi_column():
    from oilprice.fetchers import india
    values = india._parse_page_text(IN_PDF_TEXT)
    # Newest row (19-Aug), Delhi column, not Mumbai (111.21) or the
    # previous day (101.00).
    assert values["petrol"] == pytest.approx(102.12)
    assert values["diesel"] == pytest.approx(95.20)


def test_india_ignores_header_and_footer_lines():
    """Only rows carrying two dates are data rows."""
    from oilprice.fetchers import india
    for line in ("Table Posted: 19-Aug-26",
                 "Delhi Mumbai Chennai Kolkata Delhi Mumbai Chennai Kolkata",
                 "Retail Selling Price of Petrol Retail Selling Price of Diesel",
                 "(Rs./Litre) (Rs./Litre)"):
        assert india._parse_page_text(line) == {}


def test_india_builds_local_prices(monkeypatch):
    from oilprice.fetchers import india

    monkeypatch.setattr(india, "_parse_pdf",
                        lambda data: {"petrol": 102.12, "diesel": 95.20})

    def fake_get(url, **kw):
        class Resp:
            text = IN_PAGE_HTML
            content = b"%PDF-fake"
        return Resp()

    monkeypatch.setattr(india.http, "get", fake_get)
    prices = {p.product: p for p in india.fetch()}
    assert prices["petrol"].price == pytest.approx(102.12)
    assert all(p.currency == "INR" and p.country_code == "IN"
               and p.unit == "litre" for p in prices.values())
    # The reference city is recorded, since India has no single national price.
    assert "Delhi" in prices["petrol"].source


def test_slot_boundaries():
    tz = config.LOCAL_TZ
    assert pipeline.current_slot(datetime(2026, 8, 13, 9, 0, tzinfo=tz))[1] == "AM"
    assert pipeline.current_slot(datetime(2026, 8, 13, 21, 0, tzinfo=tz))[1] == "PM"
